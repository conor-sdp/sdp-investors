"""Shared helpers: taxonomy resolution, normalization, header detection."""
from __future__ import annotations

import json
import os
import re
import sqlite3
from datetime import datetime, date
from pathlib import Path
from typing import Any

import yaml
import ulid as _ulid
from rapidfuzz import fuzz, process

ROOT = Path(__file__).resolve().parent.parent
DB = ROOT / "db" / "investors.db"
SRC = ROOT.parent / "Client Network"
TAXONOMY = ROOT / "taxonomy.yml"

# Auto-load .env so every script that imports _lib gets ANTHROPIC_API_KEY etc.
# Silent if no .env file (so the pipeline still runs without it).
# override=True because the Claude Code launcher pre-seeds ANTHROPIC_API_KEY=''
# (empty) into the process env, which would otherwise block our value.
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env", override=True)
except ImportError:
    pass


# ---------------------------------------------------------------------
# Taxonomy
# ---------------------------------------------------------------------
_TAX = None

def taxonomy() -> dict:
    """Lazy-load taxonomy.yml. Returns the full dict."""
    global _TAX
    if _TAX is None:
        with open(TAXONOMY) as f:
            _TAX = yaml.safe_load(f)
    return _TAX


def picklist(name: str) -> dict[str, dict]:
    """Return the slug -> {label, aliases} dict for a named picklist."""
    return taxonomy()[name]


def resolve_slug(picklist_name: str, raw_value, fuzzy_threshold: int = 88) -> tuple[str | None, float]:
    """Resolve a raw string to a taxonomy slug.
    Returns (slug, score). slug is None if no confident match.
    score is 100 for exact/alias hit, <100 for fuzzy."""
    if raw_value is None:
        return None, 0.0
    raw = str(raw_value).strip()
    if not raw:
        return None, 0.0
    pl = picklist(picklist_name)

    # 1. exact slug match (case-insensitive)
    raw_lower = raw.lower()
    for slug in pl:
        if slug == raw_lower:
            return slug, 100.0

    # 2. exact alias match (case-insensitive)
    for slug, meta in pl.items():
        for alias in meta.get("aliases", []) or []:
            if alias.lower() == raw_lower:
                return slug, 100.0
        if meta.get("label", "").lower() == raw_lower:
            return slug, 100.0

    # 3. fuzzy match across all aliases
    candidates = []
    for slug, meta in pl.items():
        candidates.append((slug, meta.get("label", slug)))
        for alias in meta.get("aliases", []) or []:
            candidates.append((slug, alias))
    if not candidates:
        return None, 0.0
    names = [c[1] for c in candidates]
    match = process.extractOne(raw, names, scorer=fuzz.WRatio)
    if match is None:
        return None, 0.0
    matched_name, score, idx = match
    if score >= fuzzy_threshold:
        return candidates[idx][0], float(score)
    return None, float(score)


# ---------------------------------------------------------------------
# Normalization
# ---------------------------------------------------------------------
_FIRM_SUFFIX_RE = re.compile(
    r"\b(llc|l\.l\.c\.|lp|l\.p\.|llp|inc|inc\.|corp|corp\.|corporation|"
    r"co|company|capital|partners|management|mgmt|advisors|holdings|"
    r"ventures|venture|group|gp|fund|funds|limited|ltd|plc|ag|sa|s\.a\.|"
    r"gmbh|bv|n\.v\.|nv|pte|kk|family office|office|investments?)\b",
    re.IGNORECASE,
)

def norm_firm_name(name: str | None) -> str:
    """Lowercase, strip punctuation + common firm suffixes, collapse whitespace."""
    if not name:
        return ""
    s = str(name).lower().strip()
    s = re.sub(r"[®©™]", "", s)
    s = re.sub(r"[^a-z0-9\s&-]", " ", s)
    s = _FIRM_SUFFIX_RE.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def norm_company_name(name: str | None) -> str:
    """Like firm-name normalization but a touch more conservative."""
    if not name:
        return ""
    s = str(name).lower().strip()
    s = re.sub(r"[®©™]", "", s)
    s = re.sub(r"[^a-z0-9\s&-]", " ", s)
    s = re.sub(r"\b(inc|inc\.|corp|llc|ltd|co)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


_EMAIL_RE = re.compile(r"[a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}")

def norm_email(email: str | None) -> str | None:
    if not email:
        return None
    s = str(email).strip().lower()
    # extract first email if multiple
    m = _EMAIL_RE.search(s)
    return m.group(0) if m else None


def norm_url(url: str | None) -> str | None:
    if not url:
        return None
    s = str(url).strip()
    if not s:
        return None
    # Add scheme if missing
    if not re.match(r"^https?://", s, re.I):
        s = "https://" + s.lstrip("/")
    return s


def extract_domain(email_or_url: str | None) -> str | None:
    """Pull a clean lowercased registrable domain out of an email or URL."""
    if not email_or_url:
        return None
    s = str(email_or_url).strip().lower()
    if "@" in s:
        s = s.split("@", 1)[1]
    s = re.sub(r"^https?://", "", s)
    s = s.split("/", 1)[0]
    s = re.sub(r"^www\.", "", s)
    s = s.strip()
    return s or None


def parse_location(raw: str | None) -> tuple[str | None, str | None]:
    """Parse 'City | State', 'City, State', 'City, Country' into (city, country_slug)."""
    if not raw:
        return None, None
    s = str(raw).strip()
    if not s:
        return None, None
    # Try splitters
    parts = re.split(r"\s*[|,]\s*", s, maxsplit=1)
    if len(parts) == 2:
        city, region = parts[0].strip(), parts[1].strip()
        country_slug = None
        if re.fullmatch(r"[A-Z]{2}", region):
            country_slug = "us"  # 2-letter region => assume US state
        else:
            country_slug, _ = resolve_slug("geography", region, fuzzy_threshold=90)
        return city or None, country_slug
    # Single token — try as country
    country_slug, _ = resolve_slug("geography", s, fuzzy_threshold=90)
    return (s if not country_slug else None), country_slug


def title_to_seniority(title: str | None) -> str | None:
    """Map a free-text title to a seniority slug."""
    if not title:
        return None
    t = str(title).lower().strip()
    if not t:
        return None
    # Order matters — longer / more specific patterns first
    rules = [
        (r"\b(co[- ]?founder|founder)\b", "founder"),
        (r"\bchief executive\b|\bceo\b", "ceo"),
        (r"\bchief investment\b|\bcio\b", "cio"),
        (r"\bchief financial\b|\bcfo\b", "cfo"),
        (r"\bmanaging partner\b", "managing_partner"),
        (r"\bmanaging director\b|\bmd\b", "managing_director"),
        (r"\bgeneral partner\b|\bpartner\b", "partner"),
        (r"\bprincipal\b", "principal"),
        (r"\bdirector\b", "director"),
        (r"\b(senior\s+)?vice president\b|\bvp\b|\bsvp\b", "vp"),
        (r"\bassociate\b", "associate"),
        (r"\banalyst\b", "analyst"),
        (r"\brepresentative\b|\brep\b", "representative"),
    ]
    for pat, slug in rules:
        if re.search(pat, t):
            return slug
    return "other"


def parse_check_size(raw: str | None) -> tuple[float | None, float | None]:
    """Parse strings like '$50M-$150M', '20-50m', '$100m+' into (min, max) USD millions."""
    if not raw:
        return None, None
    s = str(raw).lower().replace(",", "").replace("$", "")
    # Range: 50-150m, 50m-150m, 50-150
    m = re.search(r"(\d+(?:\.\d+)?)\s*(?:m|million)?\s*[-–to]+\s*(\d+(?:\.\d+)?)\s*(?:m|million|b|billion)?", s)
    if m:
        lo, hi = float(m.group(1)), float(m.group(2))
        if "b" in s[m.end() - 2:m.end() + 2]:
            hi *= 1000
        return lo, hi
    # Single value with +
    m = re.search(r"(\d+(?:\.\d+)?)\s*(m|million|b|billion)\s*\+?", s)
    if m:
        v = float(m.group(1))
        if m.group(2).startswith("b"):
            v *= 1000
        if "+" in s:
            return v, None
        return v, v
    return None, None


def check_size_bucket(lo: float | None, hi: float | None) -> str | None:
    """Map (min, max) USDm to a check_size_bucket slug."""
    v = hi if hi is not None else lo
    if v is None:
        return None
    if v < 1:    return "micro"
    if v < 5:    return "small"
    if v < 25:   return "mid"
    if v < 100:  return "large"
    return "mega"


# ---------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------
def to_iso_date(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%b %d %Y", "%B %d %Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def to_bool(v: Any) -> int | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return 1 if v else 0
    s = str(v).strip().lower()
    if s in ("true", "yes", "y", "1", "x"):
        return 1
    if s in ("false", "no", "n", "0"):
        return 0
    return None


# ---------------------------------------------------------------------
# Name splitting
# ---------------------------------------------------------------------
def split_name(full: str | None) -> tuple[str | None, str | None, str | None]:
    """Return (first, last, full_canonical). Handles 'Last, First' and 'First M. Last'."""
    if not full:
        return None, None, None
    s = re.sub(r"\s+", " ", str(full)).strip()
    if not s:
        return None, None, None
    # "Last, First"
    if "," in s:
        parts = [p.strip() for p in s.split(",", 1)]
        if len(parts) == 2 and parts[0] and parts[1]:
            return parts[1], parts[0], f"{parts[1]} {parts[0]}"
    parts = s.split(" ")
    if len(parts) == 1:
        return parts[0], None, parts[0]
    return parts[0], " ".join(parts[1:]), s


# ---------------------------------------------------------------------
# XLSX I/O
# ---------------------------------------------------------------------
def find_header_row(ws, max_probe: int = 5, min_str_cells: int = 3) -> int:
    """1-indexed header row. Looks for row with min_str_cells string cells, no datetimes."""
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=max_probe), start=1):
        cells = [c for c in row if c is not None]
        if len(cells) < min_str_cells:
            continue
        if any(isinstance(c, (datetime, date)) for c in cells):
            continue
        if not all(isinstance(c, str) for c in cells):
            continue
        if any(len(str(c).strip()) > 80 for c in cells):
            continue
        return i
    return 1


def read_sheet(path: Path, sheet: str | None = None, header_row: int | None = None):
    """Yield rows as dicts. (headers, list_of_dicts)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sname = sheet or wb.sheetnames[0]
        ws = wb[sname]
        if header_row is None:
            header_row = find_header_row(ws)
        rows = list(ws.iter_rows(values_only=True))
        if header_row == 0:
            # No headers — use generic col_0, col_1...
            max_cols = max((len(r) for r in rows), default=0)
            headers = [f"col_{i}" for i in range(max_cols)]
            data_rows = rows
        else:
            if header_row - 1 >= len(rows):
                return [], []
            headers = [str(c).strip() if c is not None else "" for c in rows[header_row - 1]]
            while headers and headers[-1] == "":
                headers.pop()
            data_rows = rows[header_row:]
        out = []
        for r in data_rows:
            r = list(r)[: len(headers)]
            r = r + [None] * (len(headers) - len(r))
            if all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
                continue
            d = {}
            for i, h in enumerate(headers):
                key = h if h else f"col_{i}"
                d[key] = r[i]
            out.append(d)
        return headers, out
    finally:
        wb.close()


# ---------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------
def connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row
    return conn


def new_ulid() -> str:
    return str(_ulid.new())


def queue_review(conn, category: str, raw_value, picklist_name: str | None,
                 suggested: str | None, score: float, context: dict) -> None:
    conn.execute(
        """INSERT INTO review_queue
           (category, picklist_name, raw_value, suggested_value, suggestion_score, context)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (category, picklist_name, str(raw_value) if raw_value is not None else None,
         suggested, score, json.dumps(context, default=str)),
    )


def write_source(conn, source_file: str, source_sheet: str | None, source_row: int,
                 entity_type: str, entity_id: str, raw_payload: dict) -> None:
    conn.execute(
        """INSERT INTO sources
           (source_file, source_sheet, source_row, entity_type, entity_id, raw_payload)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (source_file, source_sheet, source_row, entity_type, entity_id,
         json.dumps(raw_payload, default=str)),
    )


def resolve_and_log(conn, picklist_name: str, raw, source_file: str, source_sheet: str | None,
                    source_row: int, ctx_label: str = "") -> str | None:
    """resolve_slug + queue to review_queue if no confident match. Returns slug or None."""
    slug, score = resolve_slug(picklist_name, raw)
    if slug is None and raw is not None and str(raw).strip():
        # try fuzzy with lower threshold for suggestion
        sugg, sugg_score = resolve_slug(picklist_name, raw, fuzzy_threshold=0)
        queue_review(
            conn, "out_of_picklist", raw, picklist_name,
            sugg, sugg_score,
            {"source_file": source_file, "source_sheet": source_sheet,
             "source_row": source_row, "label": ctx_label},
        )
    return slug
