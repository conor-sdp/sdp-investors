"""Peek at the first 5 rows of selected sheets, treating every row as data
(not assuming row 1 = headers). Helps diagnose sheets where row 1 is blank
or merged. Read-only."""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT.parent / "Client Network"

# (filename, sheet_name) pairs to peek at
TARGETS = [
    ("3Flash SDP Client Interest Sheet (2).xlsx", "Client Progress"),
    ("3Flash SDP Client Interest Sheet (2).xlsx", "Activity Report"),
    ("3Flash SDP Client Interest Sheet (2).xlsx", "Interest"),
    ("3Flash SDP Client Interest Sheet (2).xlsx", "HDC"),
    ("SDP MASTER Book of Business.xlsx", "Master"),
    ("SDP MASTER Book of Business.xlsx", "Capital Raise Clients"),
    ("SDP MASTER Book of Business.xlsx", "Debt deals"),
    ("SDP MASTER Book of Business.xlsx", "M&A Deals"),
    ("SDP MASTER Book of Business.xlsx", "Prospective deals"),
    ("SDP MASTER Book of Business.xlsx", "Call log"),
    ("SDP MASTER Book of Business.xlsx", "Long Term"),
    ("Infra partners.xlsx", None),  # first sheet
    ("US Family Offices 2026 - Location, Industry Focus, Stage Focus (1).xlsx", None),
    ("Cirrus v2 SDP Client Interest Sheet.xlsx", "Client Progress"),
    ("Cirrus v2 SDP Client Interest Sheet.xlsx", "Interest"),
    ("Faradyne SDP Client Interest Sheet.xlsx", "Interest"),
]

def peek(path: Path, sheet: str | None, n: int = 6) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sname = sheet or wb.sheetnames[0]
        if sname not in wb.sheetnames:
            print(f"  !! sheet {sname!r} not in {wb.sheetnames}")
            return
        ws = wb[sname]
        print(f"\n--- {path.name} :: {sname} (max_col={ws.max_column}, max_row={ws.max_row}) ---")
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = list(row)
            # Trim trailing Nones
            while cells and cells[-1] is None:
                cells.pop()
            print(f"  row {i}: {cells}")
            if i >= n:
                break
    finally:
        wb.close()

for fname, sheet in TARGETS:
    p = SRC / fname
    if not p.exists():
        print(f"!! missing: {fname}")
        continue
    try:
        peek(p, sheet)
    except Exception as e:  # noqa: BLE001
        print(f"!! error reading {fname}::{sheet}: {e}")
