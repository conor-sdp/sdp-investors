"""
Scan in-scope sheets for distinct values in categorical columns.
Cluster near-duplicates with rapidfuzz. Output: ./taxonomy_raw.json with
per-field value counts + suggested canonical groupings.

Read-only on ./Client Network/. Handles header-on-row-2 sheets.
"""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from rapidfuzz import fuzz, process

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT.parent / "Client Network"
OUT = ROOT / "taxonomy_raw.json"

# ---- In-scope sources ----
# (file, sheet, header_row_1_indexed)
INTEREST_HEADERS = {"Firm", "Name", "Email", "Stage", "Firm Type"}

INSCOPE = [
    ("3Flash SDP Client Interest Sheet (2).xlsx", "Interest"),
    ("3Flash SDP Client Interest Sheet (2).xlsx", "HDC"),
    ("Cirrus v2 SDP Client Interest Sheet.xlsx", "Interest"),
    ("Faradyne SDP Client Interest Sheet.xlsx", "Interest"),
    ("SDP MASTER Book of Business.xlsx", "Master"),
    ("Infra partners.xlsx", None),  # first sheet
    ("US Family Offices 2026 - Location, Industry Focus, Stage Focus (1).xlsx", None),
]


def find_header_row(ws, max_probe: int = 5) -> int:
    """Return 1-indexed row number that looks like the header.
    Heuristic: row with >=3 non-empty string cells and no datetime values.
    """
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=max_probe), start=1):
        cells = [c for c in row if c is not None]
        if len(cells) < 3:
            continue
        if any(hasattr(c, "isoformat") for c in cells):  # datetime row
            continue
        if all(isinstance(c, str) and len(str(c).strip()) <= 60 for c in cells):
            return i
    return 1


def read_rows(path: Path, sheet: str | None):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sname = sheet or wb.sheetnames[0]
        ws = wb[sname]
        hdr = find_header_row(ws)
        rows = list(ws.iter_rows(values_only=True))
        if hdr - 1 >= len(rows):
            return [], []
        headers = [str(c).strip() if c is not None else "" for c in rows[hdr - 1]]
        # Trim trailing empties
        while headers and headers[-1] == "":
            headers.pop()
        data = []
        for r in rows[hdr:]:
            r = list(r)[: len(headers)]
            # Pad
            r = r + [None] * (len(headers) - len(r))
            if all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
                continue
            data.append({headers[i] if headers[i] else f"col_{i}": r[i] for i in range(len(headers))})
        return headers, data
    finally:
        wb.close()


def norm(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"\s+", " ", s)
    return s


def cluster(values: Counter, min_count: int = 1, sim_threshold: int = 87):
    """Group near-duplicate values with rapidfuzz. Returns list of {canonical, members, total}."""
    items = [v for v, c in values.most_common() if c >= min_count and v]
    seen = set()
    clusters = []
    for v in items:
        if v in seen:
            continue
        group = [v]
        seen.add(v)
        # Find similar values among the rest
        for other in items:
            if other in seen:
                continue
            if fuzz.ratio(v.lower(), other.lower()) >= sim_threshold:
                group.append(other)
                seen.add(other)
        canonical = group[0]  # highest-count member
        total = sum(values[m] for m in group)
        clusters.append({
            "canonical": canonical,
            "members": group if len(group) > 1 else None,
            "count": total,
        })
    return sorted(clusters, key=lambda x: -x["count"])


def main():
    # Aggregate values by field name across sources
    field_values: dict[str, Counter] = defaultdict(Counter)
    source_summary = []

    for fname, sheet in INSCOPE:
        path = SRC / fname
        headers, data = read_rows(path, sheet)
        source_summary.append({
            "file": fname,
            "sheet": sheet or "(first)",
            "rows": len(data),
            "headers": headers,
        })
        # Track values for categorical-ish fields
        for row in data:
            for col, val in row.items():
                nv = norm(val)
                if nv is None:
                    continue
                # Only track values that look categorical (short)
                if len(nv) <= 80:
                    field_values[col][nv] += 1

    # Focus on fields we care about for taxonomy
    KEY_FIELDS = [
        "Stage", "Firm Type", "Position", "Who Replied", "Location",
        "Entity Type", "Category", "Industry Interest",
        "Industry Focus", "Stage Focus", "Type",
    ]
    out = {
        "sources": source_summary,
        "field_distinct_counts": {f: len(field_values[f]) for f in KEY_FIELDS if f in field_values},
        "fields": {},
    }
    for f in KEY_FIELDS:
        if f not in field_values:
            continue
        c = field_values[f]
        out["fields"][f] = {
            "total_distinct": len(c),
            "top_50": c.most_common(50),
            "clusters": cluster(c, min_count=1, sim_threshold=88)[:40],
        }

    OUT.write_text(json.dumps(out, indent=2, default=str))
    print(f"Wrote {OUT.relative_to(ROOT.parent)}")
    print()
    print("=" * 78)
    print("VALUE SCAN SUMMARY")
    print("=" * 78)
    for src in source_summary:
        print(f"  {src['file']} :: {src['sheet']}  ({src['rows']} rows)")
        print(f"    headers: {src['headers']}")
    print()
    print("Distinct values per key field:")
    for f, n in out["field_distinct_counts"].items():
        print(f"  {f:<22} {n:>5} distinct")
    print()
    for f, info in out["fields"].items():
        print(f"--- {f}  (distinct={info['total_distinct']}) ---")
        for entry in info["clusters"][:12]:
            marker = "" if entry["members"] is None else f"  <-- merges {entry['members']}"
            print(f"  {entry['count']:>4}  {entry['canonical']}{marker}")
        if len(info["clusters"]) > 12:
            print(f"  ... +{len(info['clusters']) - 12} more clusters")
        print()


if __name__ == "__main__":
    main()
