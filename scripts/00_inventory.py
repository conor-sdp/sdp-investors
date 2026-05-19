"""
Phase 0: Inventory every XLSX in ./Client Network/.

For each file, record: filename, sheet names, columns per sheet, row count,
file size, mtime. Cluster files by a header-fingerprint so we can see which
follow the SDP canonical layout and which are ad hoc.

Output: ./inventory.json + a human-readable summary printed to stdout.
Reads ./Client Network/ read-only.
"""
from __future__ import annotations

import json
import hashlib
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT.parent / "Client Network"
OUT = ROOT / "inventory.json"

# Canonical SDP sheet names
CANONICAL_SHEETS = {"Investor Universe", "Criteria & Legend", "Outreach Tracker"}


def header_fingerprint(columns: list[str]) -> str:
    """Stable hash of normalized column headers for clustering."""
    norm = sorted([str(c).strip().lower() for c in columns if c is not None])
    return hashlib.md5("|".join(norm).encode()).hexdigest()[:10]


def read_sheet_meta(path: Path) -> dict:
    """Return per-sheet metadata without loading full data into memory."""
    sheets = []
    # read_only=True streams rows; data_only=True returns cached values for formulas
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Pull first row for headers; count rows by streaming
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                header_row = ()
            headers = [str(c) if c is not None else "" for c in header_row]
            # Strip trailing empty headers
            while headers and headers[-1] == "":
                headers.pop()
            row_count = 0
            for _ in rows_iter:
                row_count += 1
            sheets.append({
                "sheet": sheet_name,
                "columns": headers,
                "column_count": len(headers),
                "row_count": row_count,  # excludes header
                "fingerprint": header_fingerprint(headers),
            })
    finally:
        wb.close()
    return sheets


def main():
    if not SOURCE_DIR.exists():
        raise SystemExit(f"Source folder not found: {SOURCE_DIR}")

    files_meta = []
    fp_clusters: dict[str, list[str]] = defaultdict(list)
    canonical_files: list[str] = []

    xlsx_paths = sorted(p for p in SOURCE_DIR.rglob("*.xlsx") if not p.name.startswith("~$"))
    print(f"Scanning {len(xlsx_paths)} files in {SOURCE_DIR} ...\n")

    for path in xlsx_paths:
        stat = path.stat()
        try:
            sheets = read_sheet_meta(path)
            err = None
        except Exception as e:  # noqa: BLE001
            sheets = []
            err = f"{type(e).__name__}: {e}"

        sheet_names = {s["sheet"] for s in sheets}
        is_canonical = CANONICAL_SHEETS.issubset(sheet_names)
        if is_canonical:
            canonical_files.append(path.name)

        for s in sheets:
            fp_clusters[s["fingerprint"]].append(f"{path.name} :: {s['sheet']}")

        files_meta.append({
            "file": path.name,
            "path": str(path.relative_to(ROOT.parent)),
            "size_bytes": stat.st_size,
            "mtime": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
            "sheet_count": len(sheets),
            "is_canonical": is_canonical,
            "sheets": sheets,
            "error": err,
        })

    inventory = {
        "scanned_at": datetime.now(tz=timezone.utc).isoformat(),
        "source_dir": str(SOURCE_DIR),
        "file_count": len(files_meta),
        "total_sheets": sum(f["sheet_count"] for f in files_meta),
        "total_rows": sum(s["row_count"] for f in files_meta for s in f["sheets"]),
        "canonical_files": canonical_files,
        "header_clusters": {
            fp: sheets for fp, sheets in sorted(
                fp_clusters.items(), key=lambda kv: -len(kv[1])
            )
        },
        "files": files_meta,
    }

    OUT.write_text(json.dumps(inventory, indent=2))

    # ---- Human-readable summary ----
    print("=" * 78)
    print(f"INVENTORY SUMMARY  ({inventory['scanned_at']})")
    print("=" * 78)
    print(f"Files:        {inventory['file_count']}")
    print(f"Sheets:       {inventory['total_sheets']}")
    print(f"Total rows:   {inventory['total_rows']:,}")
    print(f"Canonical:    {len(canonical_files)} / {inventory['file_count']}")
    print()
    print("Per-file breakdown:")
    print(f"  {'FILE':<60} {'SHEETS':>6} {'ROWS':>8}  CANON")
    print(f"  {'-'*60} {'-'*6} {'-'*8}  -----")
    for f in files_meta:
        rows = sum(s["row_count"] for s in f["sheets"])
        canon = "yes" if f["is_canonical"] else "no"
        name = f["file"] if len(f["file"]) <= 60 else f["file"][:57] + "..."
        print(f"  {name:<60} {f['sheet_count']:>6} {rows:>8}  {canon}")
    print()
    print("Sheet-header clusters (fingerprint -> sheets sharing identical headers):")
    for fp, sheets in inventory["header_clusters"].items():
        marker = "  *" if len(sheets) > 1 else "   "
        print(f"{marker} [{fp}]  {len(sheets)} sheet(s)")
        for s in sheets[:6]:
            print(f"        - {s}")
        if len(sheets) > 6:
            print(f"        ... +{len(sheets) - 6} more")
    print()
    print(f"Wrote {OUT.relative_to(ROOT.parent)}")


if __name__ == "__main__":
    main()
