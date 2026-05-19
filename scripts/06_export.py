"""
Phase 6: Master XLSX export.

Writes two files:
  exports/SDP_Investor_Network_v{YYYYMMDD}.xlsx  — one sheet per canonical table
  exports/review_queue_v{YYYYMMDD}.xlsx          — open review items

Formatting: navy header row, white bold text, frozen first row, banded rows,
autosized columns. The "house" SDP look.
"""
from __future__ import annotations

import json
import sqlite3
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, str(Path(__file__).parent))
from _lib import connect, ROOT


# --------- SDP brand styling ----------
NAVY = "FF0B2545"
WHITE = "FFFFFFFF"
BAND = "FFF4F6FB"
THIN = Side(style="thin", color="FFD7DEE8")

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
BAND_FILL = PatternFill("solid", fgColor=BAND)
CELL_FONT = Font(name="Calibri", size=11)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
CELL_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


# Sheets to export: (sheet_name, sql_query)
# Relationships is intentionally first — it's the primary view for proposal-building.
SHEETS = [
    ("Relationships",
     """SELECT firm, firm_type, pb_investor_type, apollo_industry,
               aum_usd_m, pb_total_investments, pb_active_portfolio,
               apollo_employee_count, apollo_founded_year,
               accepts_debt, accepts_equity, accepts_project_finance,
               accepts_credit, accepts_growth,
               extracted_sectors, extracted_geographies,
               extracted_check_min_usd_m, extracted_check_max_usd_m,
               pb_preferred_investment_types, pb_preferred_industries,
               contact_name, contact_title, contact_seniority,
               contact_email, contact_linkedin, relationship_owner,
               pb_primary_contact_name, pb_primary_contact_email,
               pb_primary_contact_title, pb_primary_contact_phone,
               firm_strategy, firm_sectors, firm_stages, check_size_bucket,
               firm_check_size_min_usd_m, firm_check_size_max_usd_m,
               hq_city, hq_country, pb_hq_location,
               apollo_hq_city, apollo_hq_state, apollo_hq_country,
               domain, url, pb_website, firm_linkedin, firm_twitter, firm_phone,
               last_sdp_client, last_engagement_date, last_engagement_status,
               last_engagement_channel, last_feedback, last_feedback_2,
               last_notes, last_responded_by,
               total_engagements_for_firm,
               mandate_count, mandate_descriptions,
               mandate_check_size_min_usd_m, mandate_check_size_max_usd_m,
               apollo_keywords, apollo_industries, mandate_signal_score,
               attio_connection_strength, attio_last_interaction,
               contact_bio, firm_aliases, enrichment_status, apollo_status,
               pb_provenance, relationship_id, firm_id, contact_id
        FROM v_relationships
        ORDER BY
          CASE WHEN last_engagement_date IS NOT NULL THEN 0 ELSE 1 END,
          last_engagement_date DESC,
          firm, contact_name"""),
    ("Firms",
     """SELECT firm_id, name_canonical, name_aliases, domain, url, type,
               check_size_bucket, check_size_min_usd_m, check_size_max_usd_m,
               aum_usd_m, fund_size_usd_m, fund_vintage,
               strategy, stages, sectors, geographies, hq_city, hq_country,
               enrichment_status, last_enriched_at, created_at
        FROM firms ORDER BY name_canonical"""),
    ("Contacts",
     """SELECT c.contact_id, f.name_canonical AS firm, c.full_name, c.first_name,
               c.last_name, c.title, c.email, c.linkedin, c.seniority, c.role,
               c.relationship_owner, c.last_contacted_at, c.bio
        FROM contacts c JOIN firms f ON f.firm_id=c.firm_id
        ORDER BY f.name_canonical, c.full_name"""),
    ("Mandates",
     """SELECT m.mandate_id, f.name_canonical AS firm, m.description, m.sectors,
               m.stages, m.check_size_min_usd_m, m.check_size_max_usd_m,
               m.geographies, m.active, m.source, m.evidence_url, m.as_of
        FROM mandates m JOIN firms f ON f.firm_id=m.firm_id
        ORDER BY f.name_canonical"""),
    ("Engagements",
     """SELECT e.engagement_id, f.name_canonical AS firm, c.full_name AS contact,
               c.email AS contact_email, e.sdp_client, e.date, e.channel, e.status,
               e.feedback, e.feedback_secondary, e.notes, e.followup, e.meeting_held,
               e.responded_by, e.smartlead_link
        FROM engagements e
        JOIN firms f ON f.firm_id=e.firm_id
        LEFT JOIN contacts c ON c.contact_id=e.contact_id
        ORDER BY e.date DESC NULLS LAST, f.name_canonical"""),
    ("Co_Investments",
     """SELECT a.name_canonical AS firm_a, b.name_canonical AS firm_b,
               ci.shared_company_count, ci.shared_companies, ci.last_updated_at
        FROM co_investments ci
        JOIN firms a ON a.firm_id=ci.a_firm_id
        JOIN firms b ON b.firm_id=ci.b_firm_id
        ORDER BY ci.shared_company_count DESC"""),
    ("Portfolio_Companies",
     """SELECT pc.pc_id, f.name_canonical AS firm, pc.company_name, pc.sector,
               pc.stage, pc.source_url
        FROM portfolio_companies pc JOIN firms f ON f.firm_id=pc.firm_id
        ORDER BY f.name_canonical, pc.company_name"""),
    ("Sources",
     """SELECT source_id, source_file, source_sheet, source_row, entity_type,
               entity_id, ingested_at
        FROM sources ORDER BY source_file, source_sheet, source_row"""),
]


def write_sheet(ws, headers, rows, sheet_name: str):
    ws.title = sheet_name[:31]  # excel sheet name max 31
    # Header
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    # Data
    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGN
            if i % 2 == 0:
                cell.fill = BAND_FILL

    # Autofit-ish — clamp to reasonable widths
    for j, h in enumerate(headers, start=1):
        col_letter = get_column_letter(j)
        # measure first 200 rows
        sample = [str(h or "")] + [
            str(ws.cell(row=r, column=j).value or "") for r in range(2, min(202, ws.max_row + 1))
        ]
        max_len = max(len(s.split("\n")[0]) for s in sample)
        width = max(10, min(60, max_len + 2))
        ws.column_dimensions[col_letter].width = width

    # Apply auto-filter on the header row
    ws.auto_filter.ref = ws.dimensions


def rows_for(conn, sql):
    rows = conn.execute(sql).fetchall()
    if not rows:
        return [], []
    headers = list(rows[0].keys())
    out = []
    for r in rows:
        out.append([_render(r[h]) for h in headers])
    return headers, out


def _render(v):
    if isinstance(v, (list, dict)):
        return json.dumps(v)
    if isinstance(v, str) and (v.startswith("[") or v.startswith("{")):
        # already JSON — flatten for readability
        try:
            parsed = json.loads(v)
            if isinstance(parsed, list):
                return ", ".join(str(x) for x in parsed)
            if isinstance(parsed, dict):
                return json.dumps(parsed)
        except (ValueError, TypeError):
            pass
    return v


def main():
    conn = connect()
    today = date.today().strftime("%Y%m%d")
    out_dir = ROOT / "exports"
    out_dir.mkdir(exist_ok=True)
    out_main = out_dir / f"SDP_Investor_Network_v{today}.xlsx"
    out_rq = out_dir / f"review_queue_v{today}.xlsx"

    # --- Master export ---
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    for sheet_name, sql in SHEETS:
        headers, data = rows_for(conn, sql)
        ws = wb.create_sheet(title=sheet_name)
        if not headers:
            ws["A1"] = "(empty)"
            continue
        write_sheet(ws, headers, data, sheet_name)
    wb.save(out_main)
    print(f"Wrote {out_main.relative_to(ROOT.parent)}")

    # --- Review queue export ---
    headers, data = rows_for(conn,
        """SELECT review_id, category, picklist_name, raw_value, suggested_value,
                  suggestion_score, context, status, created_at
           FROM review_queue WHERE status='open'
           ORDER BY category, picklist_name, raw_value""")
    wb2 = Workbook()
    wb2.remove(wb2.active)
    ws2 = wb2.create_sheet(title="Review_Queue")
    if headers:
        write_sheet(ws2, headers, data, "Review_Queue")
    else:
        ws2["A1"] = "(no open review items)"
    wb2.save(out_rq)
    print(f"Wrote {out_rq.relative_to(ROOT.parent)}")

    conn.close()
    return out_main, out_rq


if __name__ == "__main__":
    main()
